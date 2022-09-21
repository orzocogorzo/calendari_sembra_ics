from openpyxl import load_workbook
from datetime import datetime as dt
from uuid import uuid4

wb = load_workbook("data.xlsx")

calendar = wb["calendar"]
event = wb["event"]
lifecycle = wb["lifecycle"]

events = {
    event.cell(row=row, column=1).value: {
        "name": event.cell(row=row, column=2).value,
        "description": event.cell(row=row, column=3).value,
    }
    for row in range(2, 9)
}
lifecycles = {
    lifecycle.cell(row=row, column=1).value: lifecycle.cell(row=row, column=2).value
    for row in range(2, 6)
}

headers = {
    "meta": [
        "Temps de germinació o brotació (dies)",
        "Profunditat sembra(cm)",
        "Durada del cicle fins a la recol·lecció",
        "Marc de plantació",
    ],
    "months": [
        "Gener",
        "Febrer",
        "Març",
        "Abril",
        "Maig",
        "Juny",
        "Juliol",
        "Agost",
        "Setembre",
        "Octubre",
        "Novembre",
        "Desembre",
    ],
}

family_name = None
genus_name = None
genus = {}
taxonomy = {}

for row in range(4, 142):
    if calendar.cell(row=row, column=1).value:
        family_name = calendar.cell(row=row, column=1).value.strip()

    taxonomy[family_name] = taxonomy.get(family_name, [])

    if calendar.cell(row=row, column=2).value:
        genus_name = calendar.cell(row=row, column=2).value.strip()
        lifecycle = calendar.cell(row=row, column=3).value.strip()
        genus = {
            "genus": genus_name,
            "lifecycle": lifecycle,
            "events": [],
            "meta": [
                str(calendar.cell(row=row, column=col).value).strip()
                if calendar.cell(row=row, column=col).value
                else ""
                for col in range(4, 8)
            ],
        }
        taxonomy[family_name].append(genus)

    event = {}
    for col in range(8, 32):
        index = int((col - 8 - (col % 2)) / 2)
        month = headers["months"][index]
        if col % 2 != 0:
            # date = [15, month, 2020]
            date = "2022{:02d}15".format(index + 1)
        else:
            date = "2022{:02d}01".format(index + 1)
            # date = [1, month, 2020]

        if event_type := calendar.cell(row=row, column=col).value:
            event["start"] = event.get("start", date)
            event["type"] = event_type
        else:
            if event.get("start"):
                event["end"] = date
                genus["events"].append(event)
                event = {}


event_template = """BEGIN:VEVENT
UID:{uid}
DTSTAMP:{time_stamp}
DTSTART;VALUE=DATE:{start_date}
DTEND;VALUE=DATE:{end_date}
RRULE:FREQ=YEARLY;INTERVAL=1;WKST=MO
SUMMARY:[{genus}]: {event}
DESCRIPTION:<b>{genus}</b><br/>Esdeveniment: {description}<br/><br/><b>Informació</b><br/>Familia: {family}<br/>Cicle de vida: {lifecycle}<br/>Temps de germinació o brotació (dies): {meta_1}<br/>Profunditat semba (cm): {meta_2}<br/>Durada del cicle fins la recolecció: {meta_3}<br/>Marc de plantació: {meta_4}
LOCATION:Catalunya
END:VEVENT"""

icalendar = """BEGIN:VCALENDAR
VERSION:2.0
CALSCALE:GREGORIAN
PRODID:-//Can Pujades Coop//NONSGML Calendari Hort v1.0//CA
X-WR-CALNAME:Can Pujades
REFRESH-INTERVAL;VALUE=DURATION:PT4H
X-PUBLISHED-TTL:PT4H
BEGIN:VTIMEZONE
TZID:Europe/Madrid
BEGIN:DAYLIGHT
TZOFFSETFROM:+0100
TZOFFSETTO:+0200
TZNAME:CEST
DTSTART:19700329T020000
RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=-1SU
END:DAYLIGHT
BEGIN:STANDARD
TZOFFSETFROM:+0200
TZOFFSETTO:+0100
TZNAME:CET
DTSTART:19701025T030000
RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=-1SU
END:STANDARD
END:VTIMEZONE
{events}
END:VCALENDAR"""

icalendar = icalendar.format(
    events="\n".join(
        [
            event_template.format(
                family=family_name,
                genus=genus["genus"],
                lifecycle=lifecycles[genus["lifecycle"]],
                event=events[event["type"]]["name"],
                description=events[event["type"]]["description"],
                time_stamp=dt.now().isoformat(),
                start_date=event["start"],
                end_date=event["end"],
                uid=uuid4(),
                meta_1=genus["meta"][0],
                meta_2=genus["meta"][1],
                meta_3=genus["meta"][2],
                meta_4=genus["meta"][3],
            )
            for family_name in taxonomy
            for genus in taxonomy[family_name]
            for event in genus["events"]
        ]
    )
)

with open("calendari_sembra.ics", "w", encoding="utf-8") as f:
    f.write(icalendar)
